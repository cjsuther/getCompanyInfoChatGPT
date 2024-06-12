from base.base import Base
from langchain_community.document_loaders import PyPDFLoader
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain.chains import RetrievalQA
from langchain_openai import OpenAI
from langchain_openai import OpenAIEmbeddings
from openpyxl import load_workbook
import sys
from datetime import datetime
from excel import readAllData, getCompanyPrompts, getDataRow

bm = Base()

def saveDataInModel(excelFile, quarter, fieldName, value, columnNumber, rowNumberHeader):
  workbook = load_workbook(filename=excelFile)
  sheet = workbook.worksheets[1]
  #search row with fieldName
  rowNumber = 1
  for row in sheet.rows:
    if sheet.cell(rowNumber, columnNumber).value == fieldName:
      break
    rowNumber+=1
  cellNumber = 1
  for row in sheet.iter_rows(min_col=1, max_col=40, min_row=rowNumberHeader, max_row=rowNumberHeader):
    for cell in row:
      if cell.value == quarter + 'E':
        break
      cellNumber+=1
  sheet.cell(rowNumber, cellNumber).value = int(value)
  workbook.save(filename=excelFile)

def processCompany(companyName, quarter, pdfFile, excelFile):
  llm = OpenAI(openai_api_key = bm.env['OPENAI_API_KEY'], temperature=bm.env['OPENAI_TEMPERATURE'])
  embeddings = OpenAIEmbeddings(openai_api_key = bm.env['OPENAI_API_KEY'])

  loader = PyPDFLoader(pdfFile)
  documents = loader.load()

  text_splitter = CharacterTextSplitter(chunk_size=int(bm.env['OPENAI_CHUNKSIZE']), chunk_overlap=int(bm.env['OPENAI_CHUNKOVERLAP']),)
  texts = text_splitter.split_documents(documents)

  db = Chroma.from_documents(texts, embeddings)
  retriever = db.as_retriever()

  qa = RetrievalQA.from_chain_type(llm, chain_type="stuff", retriever=retriever)

  dataSheets = readAllData("data.xlsx")

  #get prompts by company
  companyPrompts = getCompanyPrompts(companyName, dataSheets)

  workbook = load_workbook(filename="data.xlsx")
  workbook.active = workbook['variables']
  sheet = workbook['variables']
  config = dataSheets[2]
  #getCompany Config
  companyConfig = next(f for f in config if f[0] == companyName)
  print(companyConfig)
  companyVariablesNames = next(f for f in config if f[0] == 'Company')
  rowNum = getDataRow(dataSheets) + 1
  i = 4
  for companyPrompt in companyPrompts[1::]: 
    dataResponse = qa.invoke(companyPrompt)
    print(dataResponse)
    valueData = dataResponse['result'].strip().replace(',', '').replace('$', '')
    sheet.cell(rowNum, i).value = valueData
    variableName = sheet.cell(1, i).value
    variablePos = next(i for i,x in enumerate(companyVariablesNames) if x == variableName)
    saveDataInModel(excelFile, quarter, companyConfig[variablePos], valueData, companyConfig[2], companyConfig[1])
    i+=1

  sheet.cell(rowNum, 1).value = datetime.today().strftime('%Y-%m-%d')
  sheet.cell(rowNum, 2).value = pdfFile
  sheet.cell(rowNum, 3).value = companyName

  workbook.save(filename="data.xlsx")
