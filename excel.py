import openpyxl
from openpyxl import load_workbook

def readAllData(path):
  dataSheet = []
  workbook = load_workbook(filename=path)
  for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    #print(f"Title = {sheet.title}")
    data = []
    for row in sheet.rows:
      dataRow = []
      for cell in row:
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            # Skip this cell
            continue
        dataRow.append(cell.value)
        #print(f"{cell.column_letter}{cell.row} = {cell.value}")
      data.append(dataRow)
    dataSheet.append(data)
  return dataSheet

def getCompanyPrompts(company, dataSheets):
  for companyPrompts in dataSheets[0]:
    if(company == companyPrompts[0]):
      return companyPrompts

def getDataRow(dataSheets):
  return len(dataSheets[1])

